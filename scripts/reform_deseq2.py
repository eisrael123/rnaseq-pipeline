#!/usr/bin/env python3
#reform_deseq2.py

import pandas as pd
from pathlib import Path
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def main():
    if len(sys.argv) != 2:
        print("Usage: python reform_deseq2.py <results_dir>")
        sys.exit(1)

    results_dir = Path(sys.argv[1])
    input_file = results_dir / "deseq2" / "deseq2_results_genes.tsv"
    output_file = results_dir / "deseq2" / "deseq2_results_genes.xlsx"

    # Read the TSV file into a DataFrame
    df = pd.read_csv(input_file, sep='\t')

    # Order the rows based on ascending padj value
    df = df.sort_values(by='padj', ascending=True)

    # Fill blank cells in DESeq2 output columns with "NA"
    deseq2_columns = ['baseMean', 'log2FoldChange', 'lfcSE', 'stat', 'pvalue', 'padj']
    df[deseq2_columns] = df[deseq2_columns].fillna('NA')

    # Create a Pandas Excel writer using openpyxl as the engine
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='DESeq2 Results')

        # Get the workbook and sheet
        workbook = writer.book
        sheet = writer.sheets['DESeq2 Results']

        # Set the font for the entire sheet
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name='Arial', size=12)

        # Format the header
        header_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Format the gene names
        gene_font = Font(name='Arial', size=12, bold=True)
        gene_alignment = Alignment(horizontal='center', vertical='center')

        for cell in sheet['A'][1:]:
            cell.font = gene_font
            cell.alignment = gene_alignment

        # Adjust column widths to fit the content
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    print(f"Excel file created at '{output_file}'")

if __name__ == "__main__":
    main()
