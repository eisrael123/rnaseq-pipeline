#!/usr/bin/env python3
#reform_sleuth.py

import pandas as pd
from pathlib import Path
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def reformat_sleuth_results(input_file, output_file):
    # Read the TSV file into a DataFrame
    df = pd.read_csv(input_file, sep='\t')

    # Order the rows based on ascending qval value
    df = df.sort_values(by='qval', ascending=True)

    # Fill blank cells with "NA"
    df = df.fillna('NA')

    # Create a Pandas Excel writer using openpyxl as the engine
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sleuth Results')

        # Get the workbook and sheet
        workbook = writer.book
        sheet = writer.sheets['Sleuth Results']

        # Set the font for the entire sheet
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name='Arial', size=12)

        # Format the header
        header_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Format the target_id column
        target_id_font = Font(name='Arial', size=12, bold=True)
        target_id_alignment = Alignment(horizontal='center', vertical='center')

        for cell in sheet['A'][1:]:
            cell.font = target_id_font
            cell.alignment = target_id_alignment

        # Adjust column widths to fit the content
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    print(f"Excel file created at '{output_file}'")

def main():
    if len(sys.argv) != 2:
        print("Usage: python reform_sleuth.py <results_dir>")
        sys.exit(1)

    results_dir = Path(sys.argv[1])

    # Define input and output file paths
    wald_input_file = results_dir / "sleuth" / "sleuth_results_wald_transcripts.tsv"
    wald_output_file = results_dir / "sleuth" / "sleuth_results_wald_transcripts.xlsx"
    lrt_input_file = results_dir / "sleuth" / "sleuth_results_lrt_transcripts.tsv"
    lrt_output_file = results_dir / "sleuth" / "sleuth_results_lrt_transcripts.xlsx"

    # Reformat and save the Wald results
    reformat_sleuth_results(wald_input_file, wald_output_file)

    # Reformat and save the LRT results
    reformat_sleuth_results(lrt_input_file, lrt_output_file)

if __name__ == "__main__":
    main()
